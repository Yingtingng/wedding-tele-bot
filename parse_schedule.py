#!/usr/bin/env python3
"""
Parse wedding schedule Excel file and extract tasks to JSON format
"""
import openpyxl
from datetime import datetime, time, timedelta
import json
import re

# Name mappings
PEOPLE = {
    'bride': ['Mars'],
    'groom': ['Daniel', 'Dan'],
    'bridesmaids': ['YT', 'Karina', 'Sandy', 'Rachel', 'Michelle'],
    'groomsmen': ['Rana', 'Robert', 'Dom', 'Sandeep', 'Ryan']
}

ALL_NAMES = PEOPLE['bride'] + PEOPLE['groom'] + PEOPLE['bridesmaids'] + PEOPLE['groomsmen']

# Column indices (1-based as per Excel)
COL_TIMING = 2
COL_BRIDE = 3
COL_GROOM = 4
COL_BRIDESMAID = 5
COL_GROOMSMEN = 6

def extract_people_from_text(text):
    """Extract names from task text"""
    if not text:
        return []

    found_people = []
    for name in ALL_NAMES:
        if name.lower() in text.lower():
            found_people.append(name)

    return list(set(found_people))

def parse_time_cell(cell_value):
    """Parse time from cell value"""
    if isinstance(cell_value, time):
        return cell_value
    elif isinstance(cell_value, datetime):
        return cell_value.time()
    elif isinstance(cell_value, str):
        try:
            return datetime.strptime(cell_value, '%H:%M:%S').time()
        except:
            try:
                return datetime.strptime(cell_value, '%H:%M').time()
            except:
                return None
    return None

def merge_consecutive_tasks(tasks):
    """Merge tasks that span multiple time slots"""
    if not tasks:
        return []

    merged = []
    current = None

    for task in tasks:
        if current is None:
            current = task.copy()
        elif (current['task'].strip() == task['task'].strip() and
              current['people'] == task['people'] and
              current['role'] == task['role']):
            # Same task continues, extend end time
            current['end_time'] = task['end_time']
        else:
            # Different task, save current and start new
            merged.append(current)
            current = task.copy()

    if current:
        merged.append(current)

    return merged

def parse_excel(file_path):
    """Parse the wedding schedule Excel file"""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Final Timeline']

    tasks = []
    wedding_date = datetime(2026, 5, 9).date()

    # Start from row 3 (after headers in row 2)
    for row_idx in range(3, sheet.max_row + 1):
        timing_cell = sheet.cell(row=row_idx, column=COL_TIMING).value

        if not timing_cell:
            continue

        start_time = parse_time_cell(timing_cell)
        if not start_time:
            continue

        # Calculate end time (5 minutes later)
        start_dt = datetime.combine(wedding_date, start_time)
        end_dt = start_dt + timedelta(minutes=5)
        end_time = end_dt.time()

        # Process each column (Bride, Groom, Bridesmaid, Groomsmen)
        columns = [
            (COL_BRIDE, 'bride', PEOPLE['bride']),
            (COL_GROOM, 'groom', PEOPLE['groom']),
            (COL_BRIDESMAID, 'bridesmaid', PEOPLE['bridesmaids']),
            (COL_GROOMSMEN, 'groomsmen', PEOPLE['groomsmen'])
        ]

        for col_idx, role, default_people in columns:
            cell_value = sheet.cell(row=row_idx, column=col_idx).value

            if cell_value and str(cell_value).strip():
                task_text = str(cell_value).strip()

                # Extract specific people mentioned in the task
                mentioned_people = extract_people_from_text(task_text)

                # If no specific people mentioned, use defaults for that role
                if not mentioned_people:
                    if role in ['bridesmaid', 'groomsmen']:
                        # For group roles, include all members
                        mentioned_people = default_people.copy()
                    else:
                        mentioned_people = default_people.copy()

                task = {
                    'start_time': start_time.strftime('%H:%M'),
                    'end_time': end_time.strftime('%H:%M'),
                    'task': task_text,
                    'role': role,
                    'people': mentioned_people
                }

                tasks.append(task)

    # Merge consecutive identical tasks
    tasks_by_role = {}
    for task in tasks:
        role = task['role']
        if role not in tasks_by_role:
            tasks_by_role[role] = []
        tasks_by_role[role].append(task)

    merged_tasks = []
    for role, role_tasks in tasks_by_role.items():
        merged_tasks.extend(merge_consecutive_tasks(role_tasks))

    # Sort by start time
    merged_tasks.sort(key=lambda x: x['start_time'])

    return merged_tasks

if __name__ == '__main__':
    excel_file = '/Users/yingting/Library/CloudStorage/OneDrive-amazon.com/dev/others/M&D WEDDING 9 MAY 26 (BRIDAL PARTY COPY).xlsx'
    tasks = parse_excel(excel_file)

    print(f"\nTotal tasks extracted: {len(tasks)}")
    print("\n" + "="*80)
    print("FIRST 20 TASKS:")
    print("="*80)

    for i, task in enumerate(tasks[:20], 1):
        people_str = ', '.join(task['people'])
        print(f"{i}. [{task['start_time']}-{task['end_time']}] {task['role']}: {task['task'][:60]}")
        print(f"   People: {people_str}\n")

    # Save to JSON
    output_file = '/Users/yingting/Library/CloudStorage/OneDrive-amazon.com/dev/others/wedding-bot/schedule.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(tasks, f, indent=2, ensure_ascii=False)

    print(f"\nSaved {len(tasks)} tasks to {output_file}")
